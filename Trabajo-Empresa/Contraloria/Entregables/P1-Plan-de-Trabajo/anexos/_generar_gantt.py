"""
Genera el Gantt vertical del Plan de Trabajo (Producto 1) en formato Excel.

Diseño:
- Filas: 35 semanas del servicio (S1 a S35).
- Columnas: 9 productos (P1 a P9), una por columna.
- Celdas coloreadas por caso atípico cuando el producto está activo en esa semana.
- Leyenda inferior: descripción de cada P# y código de colores por caso atípico.

Apto para insertar como imagen en Word vertical (portrait) sin recortes.
Salida: Gantt-P1.xlsx
"""

import os
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SUSCRIPCION = date(2026, 4, 28)
TOTAL_DIAS = 240
TOTAL_SEMANAS = (TOTAL_DIAS + 6) // 7  # 35

PRODUCTOS = [
    {"num": 1, "nombre": "Plan de Trabajo",
     "caso": "Plan de Trabajo", "inicio_dia": 1, "fin_dia": 7},
    {"num": 2, "nombre": "Caso 1 — Ingesta inicial (SSI, INFOBRAS)",
     "caso": "Caso 1: Ausencia residente/supervisor", "inicio_dia": 8, "fin_dia": 30},
    {"num": 3, "nombre": "Caso 1 — Fuentes complementarias y linaje (MEF, SIAF, MEF_ANDAT)",
     "caso": "Caso 1: Ausencia residente/supervisor", "inicio_dia": 31, "fin_dia": 60},
    {"num": 4, "nombre": "Caso 2 — Ingesta inicial (SERVIR, SUNEDU, RNP, AIRHSP)",
     "caso": "Caso 2: Sancionados SERVIR", "inicio_dia": 61, "fin_dia": 90},
    {"num": 5, "nombre": "Caso 2 — Linaje, modelo y certificación",
     "caso": "Caso 2: Sancionados SERVIR", "inicio_dia": 91, "fin_dia": 120},
    {"num": 6, "nombre": "Caso 3 — Ingesta inicial (SISFOH, JUNTOS, PENSIÓN 65)",
     "caso": "Caso 3: Programas sociales", "inicio_dia": 121, "fin_dia": 150},
    {"num": 7, "nombre": "Caso 3 — Fuentes complementarias y linaje (MEF, PLANILLAS, RENIEC)",
     "caso": "Caso 3: Programas sociales", "inicio_dia": 151, "fin_dia": 180},
    {"num": 8, "nombre": "Despliegue, matriz bus dimensional y validación P2/P5",
     "caso": "Casos 1 y 2: Despliegue y validación", "inicio_dia": 181, "fin_dia": 210},
    {"num": 9, "nombre": "Validación P6/P7, observaciones y transferencia",
     "caso": "Caso 3: Validación y transferencia", "inicio_dia": 211, "fin_dia": 240},
]

COLORES_CASO = {
    "Plan de Trabajo": "BFBFBF",
    "Caso 1: Ausencia residente/supervisor": "4472C4",
    "Caso 2: Sancionados SERVIR": "ED7D31",
    "Caso 3: Programas sociales": "70AD47",
    "Casos 1 y 2: Despliegue y validación": "FFC000",
    "Caso 3: Validación y transferencia": "7030A0",
}

# ===================== Estilos =====================

font_title = Font(name="Arial", size=13, bold=True, color="FFFFFF")
font_subtitle = Font(name="Arial", size=9, italic=True)
font_section = Font(name="Arial", size=10, bold=True, color="FFFFFF")
font_header = Font(name="Arial", size=9, bold=True, color="FFFFFF")
font_body = Font(name="Arial", size=9)
font_body_bold = Font(name="Arial", size=9, bold=True)
font_small = Font(name="Arial", size=8, italic=True)

fill_title = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
fill_section = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
fill_header = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
fill_alt = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="BFBFBF")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

# ===================== Workbook =====================

wb = Workbook()
ws = wb.active
ws.title = "Gantt P1"

NUM_COLS_GANTT = 11  # A=Semana, B=Periodo, C..K = P1..P9

# ----- Título -----
ws.cell(row=1, column=1, value="CRONOGRAMA — PLAN DE TRABAJO (PRODUCTO N° 1)")
ws.cell(row=1, column=1).font = font_title
ws.cell(row=1, column=1).fill = fill_title
ws.cell(row=1, column=1).alignment = align_center
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS_GANTT)
ws.row_dimensions[1].height = 26

ws.cell(row=2, column=1, value=(
    "Consultoría: SCI N° 013-2026-CG-UE002/BID3   |   "
    "Contrato N° 062-2026-CG-UE002/BID   |   Suscripción: 28/04/2026"
))
ws.cell(row=2, column=1).font = font_subtitle
ws.cell(row=2, column=1).alignment = align_center
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLS_GANTT)

# ----- Cabecera del Gantt -----
fila_header = 4
ws.cell(row=fila_header, column=1, value="Semana")
ws.cell(row=fila_header, column=2, value="Periodo")
for i, p in enumerate(PRODUCTOS):
    ws.cell(row=fila_header, column=3 + i, value=f"P{p['num']}")

for col in range(1, NUM_COLS_GANTT + 1):
    celda = ws.cell(row=fila_header, column=col)
    celda.font = font_header
    celda.fill = fill_header
    celda.alignment = align_center
    celda.border = border_all
ws.row_dimensions[fila_header].height = 22

# ----- Filas semanales -----
for sem in range(TOTAL_SEMANAS):
    fila = fila_header + 1 + sem
    fecha_ini = SUSCRIPCION + timedelta(days=sem * 7 + 1)  # día 1 = día siguiente a suscripción
    fecha_fin = SUSCRIPCION + timedelta(days=(sem + 1) * 7)

    ws.cell(row=fila, column=1, value=f"S{sem + 1}")
    ws.cell(row=fila, column=2,
            value=f"{fecha_ini.strftime('%d/%m')} – {fecha_fin.strftime('%d/%m')}")

    for i, p in enumerate(PRODUCTOS):
        celda = ws.cell(row=fila, column=3 + i)
        dia_ini_sem = sem * 7 + 1
        dia_fin_sem = (sem + 1) * 7
        if p["inicio_dia"] <= dia_fin_sem and p["fin_dia"] >= dia_ini_sem:
            color = COLORES_CASO[p["caso"]]
            celda.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Filas alternas con sombreado tenue (solo en columnas Semana/Periodo)
    if sem % 2 == 1:
        for col in [1, 2]:
            ws.cell(row=fila, column=col).fill = fill_alt

    # Estilo común de la fila
    for col in range(1, NUM_COLS_GANTT + 1):
        celda = ws.cell(row=fila, column=col)
        celda.font = font_body
        celda.alignment = align_center
        celda.border = border_all

# ----- Leyenda de productos -----
fila_leyenda = fila_header + 1 + TOTAL_SEMANAS + 2

ws.cell(row=fila_leyenda, column=1, value="LEYENDA DE PRODUCTOS")
ws.cell(row=fila_leyenda, column=1).font = font_section
ws.cell(row=fila_leyenda, column=1).fill = fill_section
ws.cell(row=fila_leyenda, column=1).alignment = align_center
ws.merge_cells(start_row=fila_leyenda, start_column=1,
               end_row=fila_leyenda, end_column=NUM_COLS_GANTT)
ws.row_dimensions[fila_leyenda].height = 20

# Cabecera de la leyenda
fila_leyenda_header = fila_leyenda + 1
encabezados_leyenda = ["Cód.", "Producto", "Inicio", "Fin", "Días"]
spans = [(1, 1), (2, 7), (8, 9), (10, 10), (11, 11)]
for (col_ini, col_fin), texto in zip(spans, encabezados_leyenda):
    celda = ws.cell(row=fila_leyenda_header, column=col_ini, value=texto)
    celda.font = font_header
    celda.fill = fill_header
    celda.alignment = align_center
    celda.border = border_all
    if col_ini != col_fin:
        ws.merge_cells(start_row=fila_leyenda_header, start_column=col_ini,
                       end_row=fila_leyenda_header, end_column=col_fin)
        # estilo a las celdas mergeadas
        for c in range(col_ini, col_fin + 1):
            ws.cell(row=fila_leyenda_header, column=c).border = border_all

# Filas de productos en la leyenda
for i, p in enumerate(PRODUCTOS):
    fila = fila_leyenda_header + 1 + i
    fecha_ini = SUSCRIPCION + timedelta(days=p["inicio_dia"])
    fecha_fin = SUSCRIPCION + timedelta(days=p["fin_dia"])
    duracion = p["fin_dia"] - p["inicio_dia"] + 1
    color = COLORES_CASO[p["caso"]]

    # Cód. con color de fondo
    cod = ws.cell(row=fila, column=1, value=f"P{p['num']}")
    cod.font = font_body_bold
    cod.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cod.alignment = align_center
    cod.border = border_all

    # Producto (mergeado de col 2 a 7)
    prod = ws.cell(row=fila, column=2, value=p["nombre"])
    prod.font = font_body
    prod.alignment = align_left
    prod.border = border_all
    ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=7)
    for c in range(2, 8):
        ws.cell(row=fila, column=c).border = border_all

    # Fecha inicio (mergeada col 8-9)
    ini = ws.cell(row=fila, column=8, value=fecha_ini.strftime("%d/%m/%Y"))
    ini.font = font_body
    ini.alignment = align_center
    ini.border = border_all
    ws.merge_cells(start_row=fila, start_column=8, end_row=fila, end_column=9)
    for c in range(8, 10):
        ws.cell(row=fila, column=c).border = border_all

    # Fecha fin
    fin_c = ws.cell(row=fila, column=10, value=fecha_fin.strftime("%d/%m/%Y"))
    fin_c.font = font_body
    fin_c.alignment = align_center
    fin_c.border = border_all

    # Días
    dur = ws.cell(row=fila, column=11, value=duracion)
    dur.font = font_body
    dur.alignment = align_center
    dur.border = border_all

# ----- Código de colores por caso atípico -----
fila_color = fila_leyenda_header + 1 + len(PRODUCTOS) + 2

ws.cell(row=fila_color, column=1, value="CÓDIGO DE COLOR POR CASO ATÍPICO")
ws.cell(row=fila_color, column=1).font = font_section
ws.cell(row=fila_color, column=1).fill = fill_section
ws.cell(row=fila_color, column=1).alignment = align_center
ws.merge_cells(start_row=fila_color, start_column=1,
               end_row=fila_color, end_column=NUM_COLS_GANTT)
ws.row_dimensions[fila_color].height = 20

for i, (caso, color) in enumerate(COLORES_CASO.items()):
    fila = fila_color + 1 + i
    cuadro = ws.cell(row=fila, column=1)
    cuadro.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cuadro.border = border_all

    desc = ws.cell(row=fila, column=2, value=caso)
    desc.font = font_body
    desc.alignment = align_left
    desc.border = border_all
    ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=NUM_COLS_GANTT)
    for c in range(2, NUM_COLS_GANTT + 1):
        ws.cell(row=fila, column=c).border = border_all

# ----- Nota al pie -----
fila_nota = fila_color + 1 + len(COLORES_CASO) + 1
ws.cell(row=fila_nota, column=1, value=(
    "Nota: Las fechas que coincidan con día no laborable se trasladan al día hábil "
    "siguiente, conforme al numeral 8 del TDR."
))
ws.cell(row=fila_nota, column=1).font = font_small
ws.cell(row=fila_nota, column=1).alignment = align_left
ws.merge_cells(start_row=fila_nota, start_column=1,
               end_row=fila_nota, end_column=NUM_COLS_GANTT)

# ===================== Ancho de columnas =====================

ws.column_dimensions["A"].width = 6   # Semana / Cód.
ws.column_dimensions["B"].width = 16  # Periodo / Producto-1
for c in range(3, 12):                # P1..P9 (cols C..K)
    ws.column_dimensions[get_column_letter(c)].width = 5.5

# Congelar paneles para que la cabecera y los rótulos de semana/periodo queden visibles
ws.freeze_panes = "C5"

# Configurar impresión: orientación vertical y ajuste a una página de ancho
ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_options.horizontalCentered = True
ws.page_margins.left = 0.5
ws.page_margins.right = 0.5
ws.page_margins.top = 0.5
ws.page_margins.bottom = 0.5

# ===================== Guardar =====================

salida = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Gantt-P1.xlsx")
wb.save(salida)
print(f"Excel generado: {salida}")
